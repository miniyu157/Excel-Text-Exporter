import sys
import os
import re
import json
import csv

# --- 1. 依赖库检测 ---
try:
    import openpyxl
    from openpyxl.worksheet.formula import ArrayFormula
    from openpyxl.utils import get_column_letter
    import toml
    import yaml
except ImportError as e:
    missing_library = e.name
    print(f"错误: 缺少必要的库 '{missing_library}'。")
    if missing_library == 'yaml':
        missing_library = 'pyyaml'
    print(f"请使用此命令安装: pip install {missing_library}")
    sys.exit(1)

import datetime

def load_config(config_path):
    """加载配置文件，如果文件不存在则创建并使用默认值。"""
    DEFAULT_CONFIG_CONTENT = """
# Excel-Text-Exporter

# 控制默认生成哪些文件。
# 将不需要的格式从列表中移除即可禁用。
# 可用选项: "txt", "md_plain", "md_rich", "toml", "json", "yaml", "csv"
[outputs]
default_formats = [
    "txt", 
    "md_plain", 
    "md_rich", 
    "toml", 
    "json", 
    "yaml",
    "csv"
]
# 是否将输出的JSON文件压缩为一行。默认为 false (不压缩)。
minify_json = false

# 设置所有导出文件的存放目录名。
[paths]
output_directory = "output"

# 自定义图例部分的标题。
[legends]
named_ranges = "命名区域"
reference_list = "引用列表"
expressions = "表达式"
comments = "批注"
hyperlinks = "超链接"
conditional_formatting = "条件格式"

# 自定义单元格内引用标识的前缀。
[reference_ids]
formula_prefix = "f"
comment_prefix = "c"
hyperlink_prefix = "l"
"""
    DEFAULTS = toml.loads(DEFAULT_CONFIG_CONTENT)

    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            user_config = toml.load(f)
        
        for section, keys in DEFAULTS.items():
            if section in user_config:
                for key, value in keys.items():
                    if key not in user_config[section]:
                        user_config[section][key] = value
            else:
                user_config[section] = keys
        return user_config
    except FileNotFoundError:
        print("提示: 未找到 config.toml, 已在脚本目录生成默认配置文件。")
        with open(config_path, 'w', encoding='utf-8') as f:
            f.write(DEFAULT_CONFIG_CONTENT.strip())
        return DEFAULTS

def json_default_serializer(obj):
    """为JSON序列化提供自定义的默认转换器。"""
    if isinstance(obj, (datetime.datetime, datetime.date, datetime.time)):
        return obj.isoformat()
    return str(obj)

def get_display_width(s):
    """计算字符串的显示宽度，中文字符宽度为2，英文字符为1。"""
    width = 0
    if s is None: return 0
    text = str(s)
    for char in text:
        if '\u4e00' <= char <= '\u9fa5' or '\uff00' <= char <= '\uffef':
            width += 2
        else:
            width += 1
    return width

def get_rule_details(rule, is_markdown=False):
    """
    辅助函数，用于从单个 Rule 对象中提取详细信息。
    is_markdown 参数控制输出格式。
    """
    rule_type = type(rule).__name__
    details = [f"**类型**: {rule_type}"] if is_markdown else [f"      类型: {rule_type}"]
    prefix = "" if is_markdown else "      "
    if hasattr(rule, 'formula') and rule.formula:
        formula_str = ', '.join(map(str, rule.formula))
        details.append(f"{prefix}**公式**: `{formula_str}`" if is_markdown else f"{prefix}公式: {formula_str}")
    if hasattr(rule, 'operator') and rule.operator:
        details.append(f"{prefix}**运算符**: {rule.operator}" if is_markdown else f"{prefix}运算符: {rule.operator}")
    if hasattr(rule, 'text') and rule.text:
        details.append(f"{prefix}**文本内容**: {rule.text}" if is_markdown else f"{prefix}文本内容: {rule.text}")
    if hasattr(rule, 'dxf') and rule.dxf:
        if rule.dxf.font and hasattr(rule.dxf.font, 'color') and rule.dxf.font.color:
            details.append(f"{prefix}**字体颜色**: {rule.dxf.font.color.rgb}" if is_markdown else f"{prefix}字体颜色: {rule.dxf.font.color.rgb}")
        if rule.dxf.fill and hasattr(rule.dxf.fill, 'start_color') and rule.dxf.fill.start_color:
            details.append(f"{prefix}**背景填充色**: {rule.dxf.fill.start_color.rgb}" if is_markdown else f"{prefix}背景填充色: {rule.dxf.fill.start_color.rgb}")
    return "  \n".join(details) if is_markdown else "\n".join(details)

def generate_legends(format_type, config, formulas_map, comments_map, hyperlinks_map, named_ranges_map, conditional_formatting):
    """根据指定的格式生成所有图例部分的字符串。"""
    legend_str = ""
    cfg_legends = config['legends']
    
    if named_ranges_map:
        if format_type == 'md':
            legend_str += f"\n### {cfg_legends['named_ranges']}\n"
            for name, dest in sorted(named_ranges_map.items()):
                legend_str += f"- **`{name}`**: `{dest}`\n"
        else:
            legend_str += f"\n--- {cfg_legends['named_ranges']} ---\n"
            for name, dest in sorted(named_ranges_map.items()):
                legend_str += f"{name}: {dest}\n"

    if formulas_map or comments_map or hyperlinks_map:
        if format_type == 'md':
            legend_str += f"\n### {cfg_legends['reference_list']}\n"
            if formulas_map:
                legend_str += f"\n#### {cfg_legends['expressions']}\n"
                for tag, formula in sorted(formulas_map.items(), key=lambda i: int(i[0][len(config['reference_ids']['formula_prefix'])+1:-1])):
                    legend_str += f"- **`{tag}`**: `{formula}`\n"
            if comments_map:
                legend_str += f"\n#### {cfg_legends['comments']}\n"
                for tag, text in sorted(comments_map.items(), key=lambda i: int(i[0][len(config['reference_ids']['comment_prefix'])+1:-1])):
                    legend_str += f"- **`{tag}`**: {text}\n"
            if hyperlinks_map:
                legend_str += f"\n#### {cfg_legends['hyperlinks']}\n"
                for tag, target in sorted(hyperlinks_map.items(), key=lambda i: int(i[0][len(config['reference_ids']['hyperlink_prefix'])+1:-1])):
                    legend_str += f"- **`{tag}`**: {target}\n"
        else:
            legend_str += f"\n--- {cfg_legends['reference_list']} ---\n"
            if formulas_map:
                legend_str += f"\n  {cfg_legends['expressions']}:\n"
                for tag, formula in sorted(formulas_map.items(), key=lambda i: int(i[0][len(config['reference_ids']['formula_prefix'])+1:-1])):
                    legend_str += f"  {tag}: {formula}\n"
            if comments_map:
                legend_str += f"\n  {cfg_legends['comments']}:\n"
                for tag, text in sorted(comments_map.items(), key=lambda i: int(i[0][len(config['reference_ids']['comment_prefix'])+1:-1])):
                    legend_str += f"  {tag}: {text}\n"
            if hyperlinks_map:
                legend_str += f"\n  {cfg_legends['hyperlinks']}:\n"
                for tag, target in sorted(hyperlinks_map.items(), key=lambda i: int(i[0][len(config['reference_ids']['hyperlink_prefix'])+1:-1])):
                    legend_str += f"  {tag}: {target}\n"

    if conditional_formatting:
        if format_type == 'md':
            legend_str += f"\n### {cfg_legends['conditional_formatting']}\n"
            for cf_obj in conditional_formatting:
                legend_str += f"- **作用范围**: `{cf_obj.sqref}`\n"
                for i, rule in enumerate(cf_obj.rules):
                    legend_str += f"  - **规则 #{i + 1}**\n    - {get_rule_details(rule, is_markdown=True)}\n"
        else:
            legend_str += f"\n--- {cfg_legends['conditional_formatting']} ---\n"
            for cf_obj in conditional_formatting:
                legend_str += f"  - 作用范围: {cf_obj.sqref}\n"
                for i, rule in enumerate(cf_obj.rules):
                    legend_str += f"    - 规则 #{i + 1}\n{get_rule_details(rule)}\n"
    
    return legend_str

def export_excel_to_text(file_path, config, output_dir, name_without_ext, **output_files):
    """
    将Excel文件导出为多种归档和可视化格式的文件。
    """
    try:
        wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
        wb_values = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"\n错误：无法读取Excel文件 '{file_path}'。")
        print(f"详细信息: {e}")
        sys.exit(1)

    archive_data = {'sheets': []}
    visual_txt_content = ""
    visual_md_plain_content = ""
    visual_md_rich_content = ""
    
    cfg_ids = config['reference_ids']
    sheet_names = wb_formulas.sheetnames
    
    for sheet_idx, sheet_name in enumerate(sheet_names):
        sheet_formulas = wb_formulas[sheet_name]
        sheet_values = wb_values[sheet_name]
        
        sheet_data_for_archive = {'name': sheet_name, 'named_ranges': {}, 'conditional_formatting': [], 'cells': {}}
        full_grid_data = []
        formulas_map, comments_map, hyperlinks_map, named_ranges_map = {}, {}, {}, {}
        formula_counter, comment_counter, hyperlink_counter = 1, 1, 1
        non_empty_rows, non_empty_cols = set(), set()
        
        for name, dest in wb_formulas.defined_names.items():
            if dest.localSheetId is None or dest.localSheetId == sheet_idx:
                named_ranges_map[name] = dest.attr_text
                
        for r_idx in range(1, sheet_formulas.max_row + 1):
            row_data = []
            for c_idx in range(1, sheet_formulas.max_column + 1):
                cell_formulas = sheet_formulas.cell(row=r_idx, column=c_idx)
                cell_values = sheet_values.cell(row=r_idx, column=c_idx)
                val = cell_values.value if cell_values.value is not None else ""
                tags = []
                if str(val).strip() != "" or cell_formulas.comment or cell_formulas.hyperlink or cell_formulas.data_type == 'f':
                    non_empty_rows.add(r_idx)
                    non_empty_cols.add(c_idx)
                
                if cell_formulas.data_type == 'f':
                    formula_val = cell_formulas.value
                    real_formula_to_store = None
                    if isinstance(formula_val, str) and "__xludf.DUMMYFUNCTION" in formula_val:
                        if '"COMPUTED_VALUE"' not in formula_val: real_formula_to_store = formula_val
                    elif isinstance(formula_val, ArrayFormula): real_formula_to_store = formula_val.text
                    elif isinstance(formula_val, str): real_formula_to_store = formula_val
                    if real_formula_to_store is not None:
                        tag = f"[{cfg_ids['formula_prefix']}{formula_counter}]"
                        formulas_map[tag] = real_formula_to_store
                        tags.append(tag)
                        formula_counter += 1

                if cell_formulas.comment:
                    tag = f"[{cfg_ids['comment_prefix']}{comment_counter}]"
                    comments_map[tag] = cell_formulas.comment.text
                    tags.append(tag)
                    comment_counter += 1
                
                if cell_formulas.hyperlink:
                    tag = f"[{cfg_ids['hyperlink_prefix']}{hyperlink_counter}]"
                    hyperlinks_map[tag] = cell_formulas.hyperlink.target
                    tags.append(tag)
                    hyperlink_counter += 1

                display_text = f"{val}{''.join(tags)}"
                row_data.append(display_text)
            full_grid_data.append(row_data)

        sheet_header_txt = f"工作表: {sheet_name}\n" + "-" * 40 + "\n\n"
        sheet_header_md = f"## 工作表: {sheet_name}\n\n"
        
        if not non_empty_rows or not non_empty_cols:
            if 'txt' in output_files: visual_txt_content += sheet_header_txt + "(此工作表无数据)\n\n"
            if 'md_plain' in output_files: visual_md_plain_content += sheet_header_md + "*(此工作表无数据)*\n\n"
            if 'md_rich' in output_files: visual_md_rich_content += sheet_header_md + "*(此工作表无数据)*\n\n"
            archive_data['sheets'].append({'name': sheet_name, 'data_boundary': 'empty'})
            continue

        min_r, max_r = min(non_empty_rows), max(non_empty_rows)
        min_c, max_c = min(non_empty_cols), max(non_empty_cols)
        
        sheet_data_for_archive['data_boundary'] = f"{get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{max_r}"
        sheet_data_for_archive['named_ranges'] = named_ranges_map

        for r_idx in range(min_r, max_r + 1):
            for c_idx in range(min_c, max_c + 1):
                cell_formulas = sheet_formulas.cell(row=r_idx, column=c_idx)
                cell_values = sheet_values.cell(row=r_idx, column=c_idx)
                cell_archive_data = {}
                val = cell_values.value
                
                if isinstance(val, (datetime.datetime, datetime.date, datetime.time)): cell_archive_data['value'] = val
                elif val is not None: cell_archive_data['value'] = val

                if cell_formulas.data_type == 'f':
                    formula_val = cell_formulas.value
                    real_formula_to_store = None
                    if isinstance(formula_val, str) and "__xludf.DUMMYFUNCTION" in formula_val:
                        if '"COMPUTED_VALUE"' not in formula_val: real_formula_to_store = formula_val
                    elif isinstance(formula_val, ArrayFormula): real_formula_to_store = formula_val.text
                    elif isinstance(formula_val, str): real_formula_to_store = formula_val
                    if real_formula_to_store is not None: cell_archive_data['formula'] = real_formula_to_store

                if cell_formulas.comment: cell_archive_data['comment'] = cell_formulas.comment.text
                if cell_formulas.hyperlink: cell_archive_data['hyperlink'] = cell_formulas.hyperlink.target
                if cell_archive_data: sheet_data_for_archive['cells'][cell_formulas.coordinate] = cell_archive_data

        for cf_obj in sheet_formulas.conditional_formatting:
            for rule in cf_obj.rules:
                rule_dict = {'range': str(cf_obj.sqref), 'type': rule.type}
                if hasattr(rule, 'operator') and rule.operator: rule_dict['operator'] = rule.operator
                if hasattr(rule, 'formula') and rule.formula: rule_dict['formula'] = [str(f) for f in rule.formula]
                sheet_data_for_archive['conditional_formatting'].append(rule_dict)
        archive_data['sheets'].append(sheet_data_for_archive)

        if output_files.get("csv"):
            csv_filename = os.path.join(output_dir, f"{name_without_ext}_{sheet_name}.csv")
            with open(csv_filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)
                for r_idx_csv in range(min_r, max_r + 1):
                    row_to_write = [sheet_values.cell(row=r_idx_csv, column=c_idx_csv).value for c_idx_csv in range(min_c, max_c + 1)]
                    writer.writerow(row_to_write)
            print(f"已生成: {csv_filename}")
        
        if any(f in output_files for f in ["txt", "md_plain", "md_rich"]):
            merge_info = {}
            for merged_range in sheet_formulas.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                primary_cell = (min_row, min_col)
                merge_info[primary_cell] = {'primary': True, 'colspan': max_col - min_col + 1, 'rowspan': max_row - min_row + 1}
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        if (r, c) != primary_cell: merge_info[(r, c)] = {'primary': False}

            col_widths = {c: get_display_width(get_column_letter(c)) for c in range(min_c, max_c + 1)}
            for r_idx in range(min_r, max_r + 1):
                for c_idx in range(min_c, max_c + 1):
                    cell_text = full_grid_data[r_idx - 1][c_idx - 1]
                    col_widths[c_idx] = max(col_widths.get(c_idx, 0), get_display_width(cell_text))
            
            row_header_width = len(str(max_r))
            headers_txt = [" " * (col_widths.get(c, 0) - len(get_column_letter(c))) + get_column_letter(c) for c in range(min_c, max_c + 1)]
            visual_txt_content += sheet_header_txt + " " * row_header_width + " | " + " | ".join(headers_txt) + "\n"
            separator_txt = ["-" * col_widths.get(c, 0) for c in range(min_c, max_c + 1)]
            visual_txt_content += "-" * row_header_width + "-+-" + "-+-".join(separator_txt) + "\n"
            for r_idx in range(min_r, max_r + 1):
                line_data = [full_grid_data[r_idx-1][c_idx-1] + " " * (col_widths.get(c_idx,0) - get_display_width(full_grid_data[r_idx-1][c_idx-1])) for c_idx in range(min_c, max_c + 1)]
                visual_txt_content += f"{str(r_idx).rjust(row_header_width)} | " + " | ".join(line_data) + "\n"

            headers_md = [""] + [get_column_letter(c) for c in range(min_c, max_c + 1)]
            visual_md_plain_content += sheet_header_md + "| " + " | ".join(headers_md) + " |\n"
            visual_md_plain_content += "|:" + "--:|:" + ":|".join(["--"] * (max_c - min_c + 1)) + "|\n"
            for r_idx in range(min_r, max_r + 1):
                line_data_md = [f"**{r_idx}**"]
                for c_idx in range(min_c, max_c + 1):
                    info = merge_info.get((r_idx, c_idx))
                    if info and not info['primary']: line_data_md.append("")
                    else: line_data_md.append(full_grid_data[r_idx - 1][c_idx - 1])
                visual_md_plain_content += "| " + " | ".join(line_data_md) + " |\n"
            
            visual_md_rich_content += sheet_header_md
            if not sheet_formulas.merged_cells.ranges:
                visual_md_rich_content += "| " + " | ".join(headers_md) + " |\n"
                visual_md_rich_content += "|:" + "--:|:" + ":|".join(["--"] * (max_c - min_c + 1)) + "|\n"
                for r_idx in range(min_r, max_r + 1):
                    line_data_md = [f"**{r_idx}**"] + [full_grid_data[r_idx-1][c_idx-1] for c_idx in range(min_c, max_c+1)]
                    visual_md_rich_content += "| " + " | ".join(line_data_md) + " |\n"
            else:
                visual_md_rich_content += "<table>\n  <thead>\n    <tr>\n      <th></th>\n"
                for c in range(min_c, max_c+1): visual_md_rich_content += f"      <th>{get_column_letter(c)}</th>\n"
                visual_md_rich_content += "    </tr>\n  </thead>\n  <tbody>\n"
                for r_idx in range(min_r, max_r + 1):
                    visual_md_rich_content += f"    <tr>\n      <td><b>{r_idx}</b></td>\n"
                    for c_idx in range(min_c, max_c + 1):
                        info = merge_info.get((r_idx, c_idx))
                        if info and info['primary']:
                            visual_md_rich_content += f'      <td colspan="{info["colspan"]}" rowspan="{info["rowspan"]}">{full_grid_data[r_idx - 1][c_idx - 1]}</td>\n'
                        elif info and not info['primary']: continue
                        else: visual_md_rich_content += f"      <td>{full_grid_data[r_idx - 1][c_idx - 1]}</td>\n"
                    visual_md_rich_content += "    </tr>\n"
                visual_md_rich_content += "  </tbody>\n</table>\n"

        txt_legend = generate_legends('txt', config, formulas_map, comments_map, hyperlinks_map, named_ranges_map, sheet_formulas.conditional_formatting)
        md_legend = generate_legends('md', config, formulas_map, comments_map, hyperlinks_map, named_ranges_map, sheet_formulas.conditional_formatting)
        
        if 'txt' in output_files: visual_txt_content += txt_legend
        if 'md_plain' in output_files: visual_md_plain_content += md_legend
        if 'md_rich' in output_files: visual_md_rich_content += md_legend
    
    if (path := output_files.get("toml")):
        with open(path, 'w', encoding='utf-8') as f: toml.dump(archive_data, f)
        print(f"已生成: {path}")
    if (path := output_files.get("json")):
        indent = None if config['outputs']['minify_json'] else 4
        with open(path, 'w', encoding='utf-8') as f: json.dump(archive_data, f, ensure_ascii=False, indent=indent, default=json_default_serializer)
        print(f"已生成: {path}")
    if (path := output_files.get("yaml")):
        with open(path, 'w', encoding='utf-8') as f: yaml.dump(archive_data, f, allow_unicode=True, sort_keys=False)
        print(f"已生成: {path}")
    if (path := output_files.get("txt")):
        with open(path, 'w', encoding='utf-8') as f: f.write(visual_txt_content)
        print(f"已生成: {path}")
    if (path := output_files.get("md_plain")):
        with open(path, 'w', encoding='utf-8') as f: f.write(visual_md_plain_content)
        print(f"已生成: {path}")
    if (path := output_files.get("md_rich")):
        with open(path, 'w', encoding='utf-8') as f: f.write(visual_md_rich_content)
        print(f"已生成: {path}")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("用法: python export_excel.py <example.xlsx>")
        sys.exit(1)
    input_excel_file = sys.argv[1]
    if not os.path.exists(input_excel_file):
        print(f"错误: 文件 '{input_excel_file}' 不存在。")
        sys.exit(1)
    if not input_excel_file.lower().endswith('.xlsx'):
        print(f"错误: 文件 '{input_excel_file}' 不是 .xlsx 格式。")
        sys.exit(1)

    script_dir = os.path.dirname(os.path.realpath(__file__))
    config = load_config(os.path.join(script_dir, 'config.toml'))
    output_dir = os.path.join(script_dir, config['paths']['output_directory'])
    os.makedirs(output_dir, exist_ok=True)
    
    base_name = os.path.basename(input_excel_file)
    name_without_ext = os.path.splitext(base_name)[0]

    enabled_formats = config['outputs']['default_formats']
    output_files = {}

    if 'toml' in enabled_formats: output_files['toml'] = os.path.join(output_dir, f"{name_without_ext}_archive.toml")
    if 'json' in enabled_formats: output_files['json'] = os.path.join(output_dir, f"{name_without_ext}_archive.json")
    if 'yaml' in enabled_formats: output_files['yaml'] = os.path.join(output_dir, f"{name_without_ext}_archive.yaml")
    if 'txt' in enabled_formats: output_files['txt'] = os.path.join(output_dir, f"{name_without_ext}_visual.txt")
    if 'md_plain' in enabled_formats: output_files['md_plain'] = os.path.join(output_dir, f"{name_without_ext}_visual_plain.md")
    if 'md_rich' in enabled_formats: output_files['md_rich'] = os.path.join(output_dir, f"{name_without_ext}_visual_rich.md")
    if 'csv' in enabled_formats: output_files['csv'] = True

    print(f"正在处理文件: {input_excel_file}")
    try:
        export_excel_to_text(file_path=input_excel_file,
                             config=config,
                             output_dir=output_dir,
                             name_without_ext=name_without_ext,
                             **output_files)
        print("\n处理完成！")
    except Exception as e:
        print(f"\n处理过程中发生未知错误: {e}")